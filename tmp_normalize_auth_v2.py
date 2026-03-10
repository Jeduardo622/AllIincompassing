import openpyxl, json, re
from datetime import datetime
from pathlib import Path

xlsx = r'c:\Users\test\Desktop\AllIincompassing\client authorization.xlsx'
out = Path(r'c:\Users\test\Desktop\AllIincompassing\tmp_client_authorization_normalized_v2.json')
wb = openpyxl.load_workbook(xlsx, data_only=True)
ws = wb[wb.sheetnames[0]]
num_re = re.compile(r'(\d+(?:\.\d+)?)')

def clean(s):
    if s is None:
        return None
    s = str(s).strip()
    return s if s else None

def parse_hours(s):
    s = clean(s)
    if not s: return None
    m = num_re.search(s)
    if not m: return None
    v = float(m.group(1))
    return int(v) if v.is_integer() else v

def parse_date_range(s):
    s = clean(s)
    if not s or '-' not in s: return (None, None)
    left, right = [x.strip() for x in s.split('-',1)]
    def one(v):
        parts = v.split('/')
        if len(parts)!=3: return None
        try:
            m,d,y = map(int, parts)
            if y < 100: y += 2000
            if y < 2000 or y > 2100: return None
            return datetime(y,m,d).date().isoformat()
        except Exception:
            return None
    return (one(left), one(right))

def split_name(name):
    name = clean(name)
    if not name: return (None,None,None)
    full = re.sub(r'\s+', ' ', name)
    if ',' in full:
        last, first = [p.strip() or None for p in full.split(',',1)]
        return (full, first, last)
    parts = full.split(' ')
    if len(parts)>=2: return (full, ' '.join(parts[:-1]), parts[-1])
    return (full, full, None)

def parse_service_units(s):
    s = clean(s)
    if not s: return (None,None)
    text = s.upper()
    ho = None
    for p in [r'H0032\s*HO\s*(\d+(?:\.\d+)?)\s*HRS?', r'H0032HO\s*(\d+(?:\.\d+)?)\s*HRS?', r'(\d+(?:\.\d+)?)\s*HRS?\s*HO\b', r'\bHO\s*(\d+(?:\.\d+)?)\s*HRS?']:
        m = re.search(p, text)
        if m: ho = float(m.group(1)); break
    sup = None
    m = re.search(r'H0032(?!\s*HO|HO)\s*(\d+(?:\.\d+)?)\s*HRS?', text)
    if m: sup = float(m.group(1))
    def as_int(v):
        if v is None: return None
        return int(v) if float(v).is_integer() else int(round(v))
    return (as_int(sup), as_int(ho))

rows=[]
for r in ws.iter_rows(min_row=3, values_only=True):
    if not r or not any(clean(c) for c in r):
        continue
    name_raw, auth_type, auth_amount, iehp, service_text, location, staff_needed, date_range = r[:8]
    full_name, first_name, last_name = split_name(name_raw)
    if not full_name or full_name.upper() == 'CO':
        continue
    auth_hours = parse_hours(auth_amount)
    start_date, end_date = parse_date_range(date_range)
    sup_units, pc_units = parse_service_units(service_text)
    notes = ' | '.join([x for x in [clean(iehp), clean(service_text), clean(staff_needed)] if x]) or None
    rows.append({
        'full_name': full_name,
        'first_name': first_name,
        'last_name': last_name,
        'city': clean(location),
        'authorized_hours_per_month': int(auth_hours) if isinstance(auth_hours,(int,float)) and float(auth_hours).is_integer() else None,
        'auth_units': auth_hours,
        'auth_start_date': start_date,
        'auth_end_date': end_date,
        'one_to_one_units': int(auth_hours) if clean(auth_type) and '1to1' in str(auth_type).replace(' ','').lower() and isinstance(auth_hours,(int,float)) else None,
        'supervision_units': sup_units,
        'parent_consult_units': pc_units,
        'notes': notes
    })
out.write_text(json.dumps(rows, ensure_ascii=True, separators=(',',':')), encoding='utf-8')
print('rows',len(rows))
print('short_names',[r['full_name'] for r in rows if len(r['full_name'])<4])
